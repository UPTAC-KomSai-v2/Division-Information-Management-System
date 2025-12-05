from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import datetime, date
from decimal import Decimal
from accounts.models import User
from documents.models import Document
from events.models import Event
from .models import Report, FacultyActivity
from .serializers import (
    ReportSerializer,
    ReportGenerateSerializer,
    ReportDataSerializer,
    FacultyActivitySerializer
)


class ReportViewSet(viewsets.ModelViewSet):
    """
    ViewSet for generating and managing reports.
    """
    
    queryset = Report.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ReportSerializer
    
    def get_queryset(self):
        """Filter reports by user or show all for admins"""
        user = self.request.user
        if user.role == User.Role.ADMIN:
            return Report.objects.all()
        return Report.objects.filter(generated_by=user)
    
    @action(detail=False, methods=['post'])
    def generate(self, request):
        """
        Generate a report based on criteria.
        Body: {
            "date_from": "2024-01-01",
            "date_to": "2024-12-31",
            "unit": "Meeting" (optional),
            "title": "Report Title" (optional)
        }
        """
        serializer = ReportGenerateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        date_from = serializer.validated_data['date_from']
        date_to = serializer.validated_data['date_to']
        unit_filter = serializer.validated_data.get('unit', None)
        title = serializer.validated_data.get('title', f"Report {date_from} to {date_to}")
        
        # Generate report data
        report_data = self._generate_faculty_activity_report(date_from, date_to, unit_filter)
        
        # Create report record
        report = Report.objects.create(
            report_type='FACULTY_ACTIVITY',
            title=title,
            date_from=date_from,
            date_to=date_to,
            unit_filter=unit_filter or '',
            generated_by=request.user,
            report_data=report_data
        )
        
        response_serializer = ReportSerializer(report)
        return Response({
            'report': response_serializer.data,
            'data': report_data
        }, status=status.HTTP_201_CREATED)
    
    def _generate_faculty_activity_report(self, date_from, date_to, unit_filter=None):
        """
        Generate faculty activity report data.
        This aggregates data from various sources to create the report.
        """
        # Get all faculty users
        faculty_users = User.objects.filter(role=User.Role.FACULTY, is_active=True)
        
        report_rows = []
        
        for faculty in faculty_users:
            # Get publications count (documents of type PUBLICATION uploaded by this faculty)
            publications = Document.objects.filter(
                uploaded_by=faculty,
                document_type=Document.DocumentType.PUBLICATION,
                created_at__date__gte=date_from,
                created_at__date__lte=date_to
            ).count()
            
            # Get service hours from events (events created or attended)
            # For now, we'll calculate based on event duration or use a default
            # In a full implementation, you might have a separate ServiceHours model
            events = Event.objects.filter(
                created_by=faculty,
                date__gte=date_from,
                date__lte=date_to
            )
            if unit_filter:
                events = events.filter(event_type__iexact=unit_filter.upper())
            
            # Estimate service hours (1 event = 2 hours average, or use actual duration if available)
            service_hours = Decimal(str(len(events) * 2))
            
            # Get trainings attended (count of events of type WORKSHOP or SEMINAR)
            trainings = events.filter(
                event_type__in=[Event.EventType.WORKSHOP, Event.EventType.SEMINAR]
            ).count()
            
            # Get unit/department from user profile
            unit = faculty.department or faculty.position or "N/A"
            
            report_rows.append({
                'facultyName': faculty.get_full_name(),
                'unit': unit,
                'publications': publications,
                'serviceHours': float(service_hours),
                'trainingsAttended': trainings,
                'faculty_id': faculty.id
            })
        
        return report_rows
    
    @action(detail=True, methods=['get'])
    def export(self, request, pk=None):
        """
        Export report to PDF or Excel.
        Query param: ?format=pdf or ?format=excel
        """
        report = self.get_object()
        export_format = request.query_params.get('format', 'pdf').lower()
        
        if export_format == 'pdf':
            return self._export_pdf(report)
        elif export_format == 'excel':
            return self._export_excel(report)
        else:
            return Response(
                {'error': 'Invalid format. Use ?format=pdf or ?format=excel'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _export_pdf(self, report):
        """Export report to PDF"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title = Paragraph(f"<b>{report.title}</b>", styles['Heading1'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # Report info
            info_text = f"Period: {report.date_from} to {report.date_to}<br/>"
            if report.unit_filter:
                info_text += f"Unit: {report.unit_filter}<br/>"
            info_text += f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M')}"
            info = Paragraph(info_text, styles['Normal'])
            elements.append(info)
            elements.append(Spacer(1, 12))
            
            # Data table
            data = report.report_data or []
            if data:
                # Table headers
                table_data = [['Faculty Name', 'Unit', 'Publications', 'Service Hours', 'Trainings']]
                
                # Table rows
                for row in data:
                    table_data.append([
                        row.get('facultyName', ''),
                        row.get('unit', ''),
                        str(row.get('publications', 0)),
                        str(row.get('serviceHours', 0)),
                        str(row.get('trainingsAttended', 0))
                    ])
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
            
            doc.build(elements)
            buffer.seek(0)
            
            from django.http import HttpResponse
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="report_{report.id}.pdf"'
            return response
            
        except ImportError:
            return Response(
                {'error': 'PDF generation requires reportlab. Install with: pip install reportlab'},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )
    
    def _export_excel(self, report):
        """Export report to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Title
            ws['A1'] = report.title
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            # Report info
            ws['A3'] = f"Period: {report.date_from} to {report.date_to}"
            if report.unit_filter:
                ws['A4'] = f"Unit: {report.unit_filter}"
            ws['A5'] = f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M')}"
            
            # Headers
            headers = ['Faculty Name', 'Unit', 'Publications', 'Service Hours', 'Trainings Attended']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Data rows
            data = report.report_data or []
            for row_idx, row_data in enumerate(data, start=8):
                ws.cell(row=row_idx, column=1, value=row_data.get('facultyName', ''))
                ws.cell(row=row_idx, column=2, value=row_data.get('unit', ''))
                ws.cell(row=row_idx, column=3, value=row_data.get('publications', 0))
                ws.cell(row=row_idx, column=4, value=row_data.get('serviceHours', 0))
                ws.cell(row=row_idx, column=5, value=row_data.get('trainingsAttended', 0))
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            from django.http import HttpResponse
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="report_{report.id}.xlsx"'
            return response
            
        except ImportError:
            return Response(
                {'error': 'Excel generation requires openpyxl. Install with: pip install openpyxl'},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )


class FacultyActivityViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing faculty activity records.
    """
    
    queryset = FacultyActivity.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = FacultyActivitySerializer
    
    def get_queryset(self):
        """Filter activities based on user role"""
        user = self.request.user
        
        # Admins can see all
        if user.role == User.Role.ADMIN:
            return FacultyActivity.objects.all()
        
        # Faculty can see their own
        if user.role == User.Role.FACULTY:
            return FacultyActivity.objects.filter(faculty=user)
        
        # Staff can see all
        return FacultyActivity.objects.all()

